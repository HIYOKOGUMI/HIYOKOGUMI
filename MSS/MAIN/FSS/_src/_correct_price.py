import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 最新ファイルを取得する関数
def get_latest_file(directory):
    files = [os.path.join(directory, f) for f in os.listdir(directory) if os.path.isfile(os.path.join(directory, f))]
    latest_file = max(files, key=os.path.getmtime)
    return latest_file

# ファイルパス設定
statistics_dir = '../../MSS/data/statistics'
products_dir = '../_data/f_products'
suggestion_base_dir = '../_data/f_suggestion'  # ベースの保存先ディレクトリ

# 保存先ディレクトリが存在しなければ生成
os.makedirs(suggestion_base_dir, exist_ok=True)

# 最新のstatisticsファイルとproductsファイルを取得
statistics_file = get_latest_file(statistics_dir)
output_file = get_latest_file(products_dir)

# ファイル読み込み
statistics_df = pd.read_csv(statistics_file, encoding='utf-8')
output_df = pd.read_csv(output_file, encoding='utf-8', dtype={'price': str})

# 割合列のデータを数値に変換（パーセント表示が含まれている場合も処理）
statistics_df['割合'] = pd.to_numeric(statistics_df['割合'].replace({'%': ''}, regex=True), errors='coerce') / 100

# priceカラムのカンマを除去して数値化
output_df['price'] = pd.to_numeric(output_df['price'].replace({',': ''}, regex=True), errors='coerce')

# 出力ファイルの名前を元にフォルダ名を設定
output_base_filename = os.path.splitext(os.path.basename(output_file))[0]
timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
suggestion_dir = os.path.join(suggestion_base_dir, f"suggestion_5_{output_base_filename}_{timestamp}")

# 新しい保存フォルダが存在しなければ生成
os.makedirs(suggestion_dir, exist_ok=True)

# ファイル名リストの生成
file_names = [
    os.path.join(suggestion_dir, "s.xlsx"),
    os.path.join(suggestion_dir, "ss.xlsx"),
    os.path.join(suggestion_dir, "sss.xlsx"),
    os.path.join(suggestion_dir, "ssss.xlsx"),
    os.path.join(suggestion_dir, "sssss.xlsx")
]

# デバッグ: 使用する割合を確認
percentages = statistics_df['割合'].dropna().sort_values(ascending=False)
for threshold, file_name in zip(percentages, file_names):
    print(f"Using threshold: {threshold * 100}% for file: {file_name}")

# 選別処理
condition_columns = ['新品、未使用', '未使用に近い', '目立った傷や汚れなし', 'やや傷や汚れあり', '傷や汚れあり', '全体的に状態が悪い']
remaining_df = output_df[output_df['price'] >= 17857]

# 各割合に基づいて分類と保存（最も易しい条件から順に処理）
for threshold, file_name in zip(percentages, file_names):
    # 各行に対して、状態ごとの基準値でフィルタ
    mask = remaining_df.apply(
        lambda row: row['price'] <= statistics_df[row['condition']].iloc[0] * (1 - threshold)
        if row['condition'] in condition_columns else False, axis=1
    )
    
    # 条件に合う行を抽出してファイル保存
    selected_df = remaining_df[mask]
    selected_df.to_excel(file_name, index=False)  # 一旦Excelファイルとして保存
    print(f"File saved: {file_name} with {selected_df.shape[0]} entries")
    
    # 残りのデータから抽出された行を除外
    remaining_df = remaining_df[~mask]

    # 列幅を自動調整（各列の最大文字列長に合わせる）
    wb = load_workbook(file_name)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # 列の文字番号を取得
        for cell in col:
            try:
                # セルの内容の長さを取得して、最大値を更新
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # B列とD列は個別に幅を広げる
        if col_letter == 'B':
            adjusted_width = max_length * 1.7  # B列は1.5倍の幅
        elif col_letter == 'D':
            adjusted_width = max_length * 2  # D列は2倍の幅
        else:
            adjusted_width = max_length + 2  # 他の列は通常の調整
        ws.column_dimensions[col_letter].width = adjusted_width

    wb.save(file_name)  # 調整後のファイルを保存
